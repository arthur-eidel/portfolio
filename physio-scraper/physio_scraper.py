"""
Swiss physiotherapist lead scraper.

Pulls physio practices from OpenStreetMap (and optionally Google Places),
checks each one's website, sorts them into buckets by web presence, drops the
ones that already have a modern site, and writes the rest to Excel for outreach.

  no_website      no real site (social-only counts here) -> priority
  wix_wordpress   built on a DIY builder (Wix, WordPress, Jimdo, ...)
  outdated_basic  has a site but old / thin / unreachable
  full_website    modern and complete -> dropped from the export

Run `python physio_scraper.py --help` for options.
"""

import argparse
import concurrent.futures
import datetime as dt
import json
import os
import re
import sqlite3
import sys
import time
from dataclasses import dataclass, asdict
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# --- defaults (override via CLI) ---

OUTPUT_FILE = "swiss_physio_leads.xlsx"
CACHE_FILE = "physio_cache.sqlite"

# Optional paid enrichment. Set the GOOGLE_API_KEY env var to enable it.
# A full sweep is a few thousand requests, so set a billing cap first.
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY", "")

MAX_WORKERS = 10
FETCH_TIMEOUT = 12          # per-website fetch, seconds
OVERPASS_TIMEOUT = 300      # Overpass query timeout, seconds

USER_AGENT = "physio-scraper/1.0 (+https://github.com/arthur-eidel)"

OVERPASS_ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
]

# 19 fully German-speaking cantons. Add CH-FR / CH-VS for the German parts
# of Fribourg and Valais.
GERMAN_CANTONS = [
    "CH-ZH", "CH-BE", "CH-LU", "CH-UR", "CH-SZ", "CH-OW", "CH-NW",
    "CH-GL", "CH-ZG", "CH-SO", "CH-BS", "CH-BL", "CH-SH", "CH-AR",
    "CH-AI", "CH-SG", "CH-GR", "CH-AG", "CH-TG",
]

SOCIAL_DOMAINS = {
    "facebook.com", "fb.com", "m.facebook.com", "instagram.com",
    "linkedin.com", "twitter.com", "x.com", "tiktok.com", "youtube.com",
    "linktr.ee", "google.com", "g.page", "business.site",
}

BUILDER_SIGNATURES = {
    "WordPress": ["wp-content", "wp-includes", "/wp-json", "wp-emoji",
                  'content="wordpress'],
    "Wix": ["wixstatic.com", "_wixcss", "wix-warmup-data",
            "static.parastorage.com", "wix.com"],
    "Jimdo": ["jimdo", "jimdofree.com", "jimcdn.com"],
    "Squarespace": ["squarespace.com", "static1.squarespace.com",
                    "squarespace-cdn.com"],
    "Weebly": ["weebly.com", "weeblycloud.com", "editmysite.com"],
    "GoDaddy": ["godaddysites.com", "img1.wsimg.com"],
    "Webnode": ["webnode."],
    "Site123": ["site123.me", "site123.com"],
    "Webflow": ["webflow.io", "assets.website-files.com"],
}

# outdated score >= this -> outdated_basic, else full_website
OUTDATED_THRESHOLD = 3


@dataclass
class Lead:
    business_id: str
    name: str = ""
    phone: str = ""
    email: str = ""
    owner: str = ""
    street: str = ""
    zip_code: str = ""
    city: str = ""
    canton: str = ""
    website: str = ""
    category: str = ""
    builder: str = ""
    notes: str = ""
    source: str = ""

    @property
    def location(self) -> str:
        return " ".join(b for b in (self.zip_code, self.city) if b)


# --- cache (lets the run resume if it dies halfway) ---

def init_cache(path: str):
    con = sqlite3.connect(path)
    con.execute(
        "CREATE TABLE IF NOT EXISTS leads (business_id TEXT PRIMARY KEY, data TEXT)"
    )
    con.commit()
    return con


def cache_get(con, business_id):
    row = con.execute(
        "SELECT data FROM leads WHERE business_id = ?", (business_id,)
    ).fetchone()
    return Lead(**json.loads(row[0])) if row else None


def cache_put(con, lead: Lead):
    con.execute(
        "INSERT OR REPLACE INTO leads (business_id, data) VALUES (?, ?)",
        (lead.business_id, json.dumps(asdict(lead))),
    )
    con.commit()


# --- source 1: OpenStreetMap (Overpass) ---

def build_overpass_query(cantons) -> str:
    areas = "\n".join(f'  area["ISO3166-2"="{c}"];' for c in cantons)
    return f"""
[out:json][timeout:{OVERPASS_TIMEOUT}];
(
{areas}
)->.de;
(
  nwr["healthcare"="physiotherapist"](area.de);
  nwr["amenity"="physiotherapist"](area.de);
  nwr["office"="physiotherapist"](area.de);
  nwr["healthcare"]["name"~"[Pp]hysio"](area.de);
  nwr["amenity"]["name"~"[Pp]hysio"](area.de);
);
out center tags;
"""


def fetch_overpass(cantons):
    print(f"[OSM] Querying physiotherapists across {len(cantons)} cantons ...")
    query = build_overpass_query(cantons)
    last_err = None
    for endpoint in OVERPASS_ENDPOINTS:
        for attempt in (1, 2):
            try:
                r = requests.post(
                    endpoint,
                    data={"data": query},
                    headers={"User-Agent": USER_AGENT, "Accept": "application/json"},
                    timeout=OVERPASS_TIMEOUT + 30,
                )
                r.raise_for_status()
                elements = r.json().get("elements", [])
                print(f"[OSM] {len(elements)} raw elements from {endpoint}")
                return elements
            except Exception as e:
                last_err = e
                print(f"[OSM] {endpoint} attempt {attempt} failed: {e}")
                time.sleep(3)
    raise RuntimeError(f"All Overpass endpoints failed: {last_err}")


def osm_to_lead(el) -> Lead:
    tags = el.get("tags", {})
    return Lead(
        business_id=f"osm:{el.get('type')}:{el.get('id')}",
        name=tags.get("name", "").strip(),
        phone=(tags.get("phone") or tags.get("contact:phone")
               or tags.get("contact:mobile") or "").strip(),
        email=(tags.get("email") or tags.get("contact:email") or "").strip(),
        owner=(tags.get("operator") or "").strip(),
        street=" ".join(x for x in (tags.get("addr:street", ""),
                                    tags.get("addr:housenumber", "")) if x).strip(),
        zip_code=tags.get("addr:postcode", "").strip(),
        city=tags.get("addr:city", "").strip(),
        canton=tags.get("addr:state", "").strip(),
        website=(tags.get("website") or tags.get("contact:website") or "").strip(),
        source="OSM",
    )


# --- source 2: Google Places (optional, paid) ---

CH_LAT_RANGE = (46.30, 47.81)   # cuts off Ticino in the south
CH_LNG_RANGE = (6.80, 10.49)    # cuts off western (French) Switzerland
GRID_STEP_DEG = 0.18            # ~13-20 km between grid points
PLACES_RADIUS = 13000           # metres
PLACES_QUERY_TERM = "Physiotherapie"


def _grid_points():
    lat = CH_LAT_RANGE[0]
    while lat <= CH_LAT_RANGE[1]:
        lng = CH_LNG_RANGE[0]
        while lng <= CH_LNG_RANGE[1]:
            yield round(lat, 4), round(lng, 4)
            lng += GRID_STEP_DEG
        lat += GRID_STEP_DEG


def fetch_google_places(session):
    if not GOOGLE_API_KEY:
        return []
    print("[Places] Sweeping via Google Places (paid) ...")
    found = {}
    base = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    for lat, lng in _grid_points():
        params = {
            "query": PLACES_QUERY_TERM,
            "location": f"{lat},{lng}",
            "radius": PLACES_RADIUS,
            "key": GOOGLE_API_KEY,
        }
        while True:
            try:
                data = session.get(base, params=params, timeout=20).json()
            except Exception as e:
                print(f"[Places] grid {lat},{lng} error: {e}")
                break
            for res in data.get("results", []):
                pid = res.get("place_id")
                if pid and pid not in found:
                    found[pid] = res
            token = data.get("next_page_token")
            if not token:
                break
            time.sleep(2)  # token isn't valid immediately
            params = {"pagetoken": token, "key": GOOGLE_API_KEY}
    print(f"[Places] {len(found)} unique places, fetching details ...")

    leads = []
    detail_url = "https://maps.googleapis.com/maps/api/place/details/json"
    for pid, res in found.items():
        try:
            d = session.get(detail_url, params={
                "place_id": pid,
                "fields": "name,formatted_phone_number,website,"
                          "formatted_address,address_components",
                "key": GOOGLE_API_KEY,
            }, timeout=20).json().get("result", {})
        except Exception:
            d = {}
        zip_code, city = "", ""
        for comp in d.get("address_components", []):
            types = comp.get("types", [])
            if "postal_code" in types:
                zip_code = comp.get("long_name", "")
            if "locality" in types:
                city = comp.get("long_name", "")
        leads.append(Lead(
            business_id=f"gplaces:{pid}",
            name=d.get("name") or res.get("name", ""),
            phone=d.get("formatted_phone_number", ""),
            website=(d.get("website") or "").strip(),
            zip_code=zip_code,
            city=city or res.get("formatted_address", ""),
            source="GooglePlaces",
        ))
    return leads


# --- website analysis ---

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")
BAD_EMAIL_HINTS = ("example.", "sentry", ".png", ".jpg", ".gif", "@2x",
                   "wixpress", "no-reply", "noreply")
YEAR_RE = re.compile(r"(?:©|&copy;|copyright)[^\d]{0,15}(20\d{2})", re.I)


def normalise_url(url: str) -> str:
    url = url.strip()
    if url and not url.startswith(("http://", "https://")):
        url = "https://" + url
    return url


def is_social_only(url: str) -> bool:
    host = urlparse(normalise_url(url)).netloc.lower().replace("www.", "")
    return any(host == d or host.endswith("." + d) for d in SOCIAL_DOMAINS)


def fetch_page(url, session):
    try:
        return session.get(url, timeout=FETCH_TIMEOUT, allow_redirects=True)
    except Exception:
        return None


def detect_builder(html_lower, headers):
    blob = html_lower + " " + " ".join(f"{k}:{v}" for k, v in headers.items()).lower()
    for builder, sigs in BUILDER_SIGNATURES.items():
        if any(sig in blob for sig in sigs):
            return builder
    return ""


def extract_email(html, soup):
    for a in soup.find_all("a", href=True):
        if a["href"].lower().startswith("mailto:"):
            addr = a["href"][7:].split("?")[0].strip()
            if addr and not any(b in addr.lower() for b in BAD_EMAIL_HINTS):
                return addr
    for m in EMAIL_RE.findall(html):
        if not any(b in m.lower() for b in BAD_EMAIL_HINTS):
            return m
    return ""


OWNER_KEYWORDS = ["inhaber", "inhaberin", "geschäftsführer", "geschäftsführerin",
                  "praxisinhaber", "praxisleitung", "verantwortlich", "leitung",
                  "propriétaire", "responsable", "titolare"]
NAME_GUESS_RE = re.compile(r"\b([A-ZÄÖÜ][a-zäöüé]+\s+[A-ZÄÖÜ][a-zäöüé]+)\b")


def extract_owner(text):
    low = text.lower()
    for kw in OWNER_KEYWORDS:
        idx = low.find(kw)
        if idx != -1:
            m = NAME_GUESS_RE.search(text[idx: idx + 120])
            if m:
                return m.group(1)
    return ""


def score_outdated(soup, final_url, html_lower):
    score = 0
    notes = []

    if not soup.find("meta", attrs={"name": "viewport"}):
        score += 2
        notes.append("not-responsive")
    if urlparse(final_url).scheme == "http":
        score += 1
        notes.append("no-https")
    if soup.find("frameset") or soup.find("frame"):
        score += 3
        notes.append("frames")
    if soup.find("font") or soup.find("center") or soup.find("marquee"):
        score += 2
        notes.append("legacy-tags")
    if ".swf" in html_lower or "shockwave-flash" in html_lower:
        score += 3
        notes.append("flash")
    if soup.find("table") and not soup.find(["section", "article", "nav", "header"]):
        score += 1
        notes.append("table-layout")

    m = YEAR_RE.search(html_lower)
    if m:
        try:
            yr = int(m.group(1))
            if yr <= dt.date.today().year - 5:
                score += 1
                notes.append(f"old-copyright-{yr}")
        except ValueError:
            pass

    if len(soup.get_text(strip=True)) < 600:
        score += 1
        notes.append("thin-content")

    return score, notes


def classify_website(lead: Lead, session) -> Lead:
    raw = lead.website.strip()

    if not raw:
        lead.category = "no_website"
        return lead

    if is_social_only(raw):
        lead.category = "no_website"
        lead.notes = (lead.notes + "; social_only: " + raw).strip("; ")
        lead.website = ""
        return lead

    resp = fetch_page(normalise_url(raw), session)
    if resp is None or resp.status_code >= 400 or not resp.text.strip():
        lead.category = "outdated_basic"
        lead.notes = (lead.notes + "; website_unreachable").strip("; ")
        return lead

    html = resp.text
    html_lower = html.lower()
    soup = BeautifulSoup(html, "html.parser")

    if not lead.email:
        lead.email = extract_email(html, soup)
    if not lead.owner:
        lead.owner = extract_owner(soup.get_text(" "))

    builder = detect_builder(html_lower, dict(resp.headers))
    if builder:
        lead.category = "wix_wordpress"
        lead.builder = builder
        return lead

    score, notes = score_outdated(soup, resp.url, html_lower)
    if notes:
        lead.notes = (lead.notes + "; " + ", ".join(notes)).strip("; ")
    lead.category = "outdated_basic" if score >= OUTDATED_THRESHOLD else "full_website"
    return lead


# --- dedup ---

def dedup(leads):
    seen = {}
    out = []
    for l in leads:
        if not l.name:
            continue
        key = (re.sub(r"\W+", "", l.name.lower()), l.zip_code)
        if key in seen:
            existing = seen[key]
            if not existing.phone and l.phone:
                existing.phone = l.phone
            if not existing.website and l.website:
                existing.website = l.website
            continue
        seen[key] = l
        out.append(l)
    return out


# --- excel output ---

CATEGORY_ORDER = {"no_website": 0, "wix_wordpress": 1, "outdated_basic": 2}
CATEGORY_COLORS = {
    "no_website": "C6EFCE",      # green = priority
    "wix_wordpress": "FFEB9C",   # yellow
    "outdated_basic": "FCE4D6",  # orange
}
HEADERS = ["Business Name", "Category", "Phone", "Email", "Owner",
           "Street", "ZIP", "City", "Canton", "Website", "Builder",
           "Notes", "Source"]


def write_excel(leads, output_file):
    keep = [l for l in leads if l.category in CATEGORY_ORDER]
    keep.sort(key=lambda l: (CATEGORY_ORDER[l.category], l.city, l.name))

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    for r, l in enumerate(keep, start=2):
        row = [l.name, l.category, l.phone, l.email, l.owner, l.street,
               l.zip_code, l.city, l.canton, l.website, l.builder,
               l.notes, l.source]
        for col, val in enumerate(row, 1):
            ws.cell(row=r, column=col, value=val)
        fill = CATEGORY_COLORS.get(l.category)
        if fill:
            ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=fill)

    widths = [34, 15, 18, 30, 22, 28, 8, 18, 10, 36, 12, 40, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(keep) + 1}"

    s = wb.create_sheet("Summary")
    counts = {}
    for l in leads:
        counts[l.category] = counts.get(l.category, 0) + 1
    s["A1"] = "Category"
    s["B1"] = "Count"
    s["A1"].font = s["B1"].font = Font(bold=True)
    order = ["no_website", "wix_wordpress", "outdated_basic", "full_website"]
    for i, cat in enumerate(order, start=2):
        s.cell(row=i, column=1, value=cat)
        s.cell(row=i, column=2, value=counts.get(cat, 0))
    s.cell(row=len(order) + 3, column=1, value="Kept (exported)")
    s.cell(row=len(order) + 3, column=2, value=len(keep))
    s.cell(row=len(order) + 4, column=1, value="Generated")
    s.cell(row=len(order) + 4, column=2, value=dt.datetime.now().strftime("%Y-%m-%d %H:%M"))
    s.column_dimensions["A"].width = 22
    s.column_dimensions["B"].width = 12

    wb.save(output_file)
    print(f"\n[OK] Wrote {len(keep)} leads to {output_file}")
    print(f"     no_website     : {counts.get('no_website', 0)}  (priority)")
    print(f"     wix_wordpress  : {counts.get('wix_wordpress', 0)}")
    print(f"     outdated_basic : {counts.get('outdated_basic', 0)}")
    print(f"     full_website   : {counts.get('full_website', 0)}  (dropped)")


# --- main ---

def parse_args():
    p = argparse.ArgumentParser(description="Swiss physiotherapist lead scraper.")
    p.add_argument("-o", "--output", default=OUTPUT_FILE,
                   help=f"Excel output file (default: {OUTPUT_FILE})")
    p.add_argument("--cache", default=CACHE_FILE,
                   help=f"SQLite cache file (default: {CACHE_FILE})")
    p.add_argument("--fresh", action="store_true",
                   help="Ignore the cache and re-check every website")
    p.add_argument("--workers", type=int, default=MAX_WORKERS,
                   help=f"Concurrent website checks (default: {MAX_WORKERS})")
    p.add_argument("--cantons", nargs="+", default=GERMAN_CANTONS, metavar="CH-XX",
                   help="ISO3166-2 canton codes to include")
    return p.parse_args()


def main():
    args = parse_args()
    con = init_cache(args.cache)

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    # 1. gather raw leads
    raw = [osm_to_lead(el) for el in fetch_overpass(args.cantons)]
    raw = [l for l in raw if l.name]
    raw.extend(fetch_google_places(session))  # no-op without a key

    leads = dedup(raw)
    print(f"[INFO] {len(leads)} unique practices to classify.")

    # 2. classify (parallel), reusing cache unless --fresh
    to_do = []
    for l in leads:
        cached = None if args.fresh else cache_get(con, l.business_id)
        if cached and cached.category:
            l.__dict__.update(asdict(cached))
        else:
            to_do.append(l)
    print(f"[INFO] {len(leads) - len(to_do)} cached, {len(to_do)} to check.")

    done = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=args.workers) as ex:
        futures = {ex.submit(classify_website, l, session): l for l in to_do}
        for fut in concurrent.futures.as_completed(futures):
            l = futures[fut]
            try:
                fut.result()
            except Exception as e:
                l.category = "outdated_basic"
                l.notes = (l.notes + f"; error:{e}").strip("; ")
            cache_put(con, l)
            done += 1
            if done % 25 == 0 or done == len(to_do):
                print(f"  ... classified {done}/{len(to_do)}")

    # 3. export
    write_excel(leads, args.output)
    con.close()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nStopped. Run again to resume from the cache.")
        sys.exit(1)
